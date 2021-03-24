import openpyxl


class HomePageData:

    test_HomePage_data = [{"firstname":"Test","lastname":"teams","gender":"Male"},{"firstname":"ticvic","lastname":"techie","gender":"Female"}]

    @staticmethod
    def getTestData(test_case_name):
        Dict = {}
        book = openpyxl.load_workbook("/home/ticvictech/Downloads/excelDemo.xlsx")
        sheet = book.active
        for i in range(1,sheet.max_row + 1):
            #print("i value ", i)
            if sheet.cell(row=i, column=1).value == test_case_name:

                for j in range(2,sheet.max_column + 1):
                    #print("j value", j)
                    #Dict["lastname"] = "teams"
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
                    #print(Dict[sheet.cell(row=1, column=j).value])
        return[Dict]